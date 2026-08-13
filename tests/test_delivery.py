"""Tests for the GT-conformant .xlsx delivery writer."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from ga_data_center_tracker.delivery import (
    Dataset,
    Variable,
    build_long_rows,
    build_transformed_rows,
    write_workbook,
)


def _sample_dataset() -> Dataset:
    county_values = {
        "Fulton County, Georgia": {"dc_operational_n": 3},
        "DeKalb County, Georgia": {"dc_operational_n": 1},
    }
    long_rows = build_long_rows(county_values)
    variables = [
        Variable(
            varname="dc_operational_n",
            definition="Operational data centers in the county.",
            units="facilities",
            source="EPA FRS",
        )
    ]
    original = [
        {"name": "Example DC", "county": "Fulton County, Georgia", "registry_id": "1"},
    ]
    return Dataset(original_rows=original, long_rows=long_rows, variables=variables)


def test_build_long_rows_shape():
    rows = build_long_rows({"Fulton County, Georgia": {"a": 1, "b": 2}})
    assert rows == [
        {"county": "Fulton County, Georgia", "varname": "a", "datavalue": 1},
        {"county": "Fulton County, Georgia", "varname": "b", "datavalue": 2},
    ]


def test_long_rows_follow_reference_order():
    # DeKalb (13089) sorts before Fulton (13121) in the reference table.
    rows = build_long_rows(
        {
            "Fulton County, Georgia": {"a": 1},
            "DeKalb County, Georgia": {"a": 2},
        }
    )
    counties = [r["county"] for r in rows]
    assert counties.index("DeKalb County, Georgia") < counties.index("Fulton County, Georgia")


def test_write_workbook_has_required_sheets(tmp_path):
    out = write_workbook(_sample_dataset(), tmp_path / "out.xlsx")
    wb = load_workbook(out)
    # Transformed is omitted when no transformed rows are supplied.
    assert wb.sheetnames == ["Original", "Long", "Data Description", "Codebook"]


def test_long_sheet_columns(tmp_path):
    out = write_workbook(_sample_dataset(), tmp_path / "out.xlsx")
    wb = load_workbook(out)
    header = [c.value for c in wb["Long"][1]]
    assert header == ["county", "varname", "datavalue"]


def test_codebook_one_row_per_variable(tmp_path):
    out = write_workbook(_sample_dataset(), tmp_path / "out.xlsx")
    wb = load_workbook(out)
    rows = list(wb["Codebook"].iter_rows(values_only=True))
    assert rows[0] == ("variable", "definition", "units")
    assert rows[1][0] == "dc_operational_n"


def test_transformed_sheet_included_when_present(tmp_path):
    ds = _sample_dataset()
    ds.transformed_rows = [{"county": "Fulton County, Georgia", "dc_operational_n": 3}]
    out = write_workbook(ds, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    assert "Transformed" in wb.sheetnames


def test_unknown_county_rejected(tmp_path):
    ds = _sample_dataset()
    ds.long_rows = [{"county": "Fulton, GA", "varname": "x", "datavalue": 1}]
    with pytest.raises(ValueError, match="reference table"):
        write_workbook(ds, tmp_path / "bad.xlsx")


def test_empty_variables_rejected(tmp_path):
    ds = _sample_dataset()
    ds.variables = []
    with pytest.raises(ValueError, match="no variables"):
        write_workbook(ds, tmp_path / "bad.xlsx")


# --- build_transformed_rows ---------------------------------------------------

def test_transformed_rows_cover_all_counties():
    rows = build_transformed_rows({"Fulton County, Georgia": {"dc_mapped_n": 45}})
    assert len(rows) == 1
    assert rows[0]["county"] == "Fulton County, Georgia"


def test_transformed_rows_follow_reference_county_order():
    rows = build_transformed_rows(
        {
            "Fulton County, Georgia": {"dc_mapped_n": 45},
            "Appling County, Georgia": {"dc_mapped_n": 1},
        }
    )
    # Reference order is alphabetical by county name, not insertion order.
    assert [r["county"] for r in rows] == [
        "Appling County, Georgia",
        "Fulton County, Georgia",
    ]


def test_transformed_rows_distinguish_missing_from_zero():
    # A variable never computed for a county is blank; a computed zero is 0.
    rows = build_transformed_rows(
        {
            "Appling County, Georgia": {"dc_mapped_n": 0, "dc_institutional_n": 0},
            "Fulton County, Georgia": {"dc_mapped_n": 45},
        }
    )
    by_county = {r["county"]: r for r in rows}
    assert by_county["Appling County, Georgia"]["dc_institutional_n"] == 0
    assert by_county["Fulton County, Georgia"]["dc_institutional_n"] == ""


def test_transformed_columns_follow_variable_registration_order():
    rows = build_transformed_rows(
        {"Fulton County, Georgia": {"dc_permitted_n": 15, "dc_mapped_n": 45}}
    )
    assert list(rows[0].keys()) == ["county", "dc_permitted_n", "dc_mapped_n"]
