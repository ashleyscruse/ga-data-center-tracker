"""Solutions Tracker delivery format.

Georgia Tech requires each dataset as an ``.xlsx`` workbook with a fixed set of
sheets (see ``docs/output_format.md``). This module builds that workbook so
delivery is mechanical rather than hand-assembled in Excel.

Sheets produced:
  - Original          the source data as pulled
  - Transformed       (only if transformed rows are supplied)
  - Long              exactly three columns: county, varname, datavalue
  - Data Description   provenance per variable
  - Codebook          one row per variable: name, definition, units

All county values must already be in the tracker form ``X County, Georgia``
(use ``counties.normalize_county``); this module validates that and refuses to
write an unknown county rather than silently shipping a bad join key.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font

from .counties import load_reference

LONG_COLUMNS = ["county", "varname", "datavalue"]


@dataclass
class Variable:
    """One variable in the dataset, with everything the Codebook and Data
    Description sheets need to document it."""

    varname: str
    definition: str
    units: str = ""
    source: str = ""
    vintage: str = ""
    date_pulled: str = ""
    original_name: str = ""
    transformations: str = ""


@dataclass
class Dataset:
    """A single deliverable: the original data, the long-format reshape, the
    variables that describe it, and optional transformed data."""

    original_rows: Sequence[Mapping[str, object]]
    long_rows: Sequence[Mapping[str, object]]
    variables: Sequence[Variable]
    transformed_rows: Sequence[Mapping[str, object]] | None = None
    notes: str = ""


def build_long_rows(
    county_values: Mapping[str, Mapping[str, object]],
) -> list[dict[str, object]]:
    """Turn ``{county: {varname: value}}`` into long rows (county, varname, datavalue).

    Counties must already be in tracker form. Order follows the reference table,
    then variable insertion order, so output is deterministic.
    """
    rows: list[dict[str, object]] = []
    # Deterministic order: reference-table county order, then var order as given.
    ref_order = [c.tracker_name for c in load_reference()]
    ordered_counties = [c for c in ref_order if c in county_values]
    # Include any counties not in the reference table at the end (will fail validation later).
    ordered_counties += [c for c in county_values if c not in ref_order]
    for county in ordered_counties:
        for varname, value in county_values[county].items():
            rows.append({"county": county, "varname": varname, "datavalue": value})
    return rows


def _valid_counties() -> set[str]:
    return {c.tracker_name for c in load_reference()}


def _validate_counties(long_rows: Iterable[Mapping[str, object]]) -> None:
    valid = _valid_counties()
    unknown = sorted(
        {str(r["county"]) for r in long_rows if str(r["county"]) not in valid}
    )
    if unknown:
        raise ValueError(
            "Long sheet contains counties not in the reference table "
            f"(must be 'X County, Georgia'): {unknown[:10]}"
            + (" ..." if len(unknown) > 10 else "")
        )


def _write_table(ws, rows: Sequence[Mapping[str, object]], columns: Sequence[str]) -> None:
    bold = Font(bold=True)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))


def write_workbook(dataset: Dataset, path: str | Path) -> Path:
    """Write ``dataset`` to a GT-conformant ``.xlsx`` workbook at ``path``."""
    if not dataset.variables:
        raise ValueError("Dataset has no variables; Codebook would be empty.")
    _validate_counties(dataset.long_rows)

    wb = Workbook()

    # Original
    ws_original = wb.active
    ws_original.title = "Original"
    original_cols = list(dataset.original_rows[0].keys()) if dataset.original_rows else []
    _write_table(ws_original, dataset.original_rows, original_cols)

    # Transformed (only if supplied)
    if dataset.transformed_rows:
        ws_t = wb.create_sheet("Transformed")
        t_cols = list(dataset.transformed_rows[0].keys())
        _write_table(ws_t, dataset.transformed_rows, t_cols)

    # Long
    ws_long = wb.create_sheet("Long")
    _write_table(ws_long, dataset.long_rows, LONG_COLUMNS)

    # Data Description
    ws_desc = wb.create_sheet("Data Description")
    desc_cols = [
        "varname",
        "source",
        "vintage",
        "date_pulled",
        "original_name",
        "definition",
        "transformations",
    ]
    desc_rows = [
        {
            "varname": v.varname,
            "source": v.source,
            "vintage": v.vintage,
            "date_pulled": v.date_pulled,
            "original_name": v.original_name,
            "definition": v.definition,
            "transformations": v.transformations,
        }
        for v in dataset.variables
    ]
    _write_table(ws_desc, desc_rows, desc_cols)

    # Codebook
    ws_code = wb.create_sheet("Codebook")
    code_rows = [
        {"variable": v.varname, "definition": v.definition, "units": v.units}
        for v in dataset.variables
    ]
    _write_table(ws_code, code_rows, ["variable", "definition", "units"])

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
