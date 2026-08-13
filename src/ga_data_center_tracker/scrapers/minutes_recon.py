"""Recon pass for county commission and zoning minutes.

The community-engagement strand needs what residents actually said, and that lives
in county commission and zoning board minutes. Georgia has 159 counties publishing
on no common system, which is why this source has been the project's hardest.

**This module does not scrape minutes. It finds out where they are.** That is the
step that makes the scrape tractable: Georgia counties do not each roll their own
agenda system, they buy one of about a dozen. Once each county is mapped to its
platform, the work collapses from 159 bespoke scrapers to a handful of adapters,
and the recon output says exactly how many counties each adapter would buy.

Two-stage detection, because a county's platform is often not visible on its
front page:

1. **Find the county's site.** There is no machine-readable directory of Georgia
   county websites, so candidate hostnames are probed directly. The patterns come
   from what Georgia counties actually use (``douglascountyga.gov``,
   ``coweta.ga.us``, ``spaldingcounty.com``), and a county that resolves to none
   of them is reported as unresolved rather than guessed at.
2. **Fingerprint the platform.** Vendor names leak into markup, script sources,
   and redirect URLs. The homepage is checked first, then any page it links to
   whose text mentions agendas or minutes, which is where the vendor usually sits.

Scope: the counties that matter, not all 159. A county with no data center and no
local action has nothing for this dataset to find, so the target set is the union
of counties with a facility and counties with a recorded ordinance or moratorium.

Run:  ``python -m ga_data_center_tracker.scrapers.minutes_recon``
"""

from __future__ import annotations

import argparse
import re
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests

USER_AGENT = "ga-data-center-tracker/0.1 (research)"

# Hostname patterns Georgia counties actually use, most common first. Probed in
# order; the first that answers with a real page wins.
HOST_PATTERNS = (
    "www.{s}countyga.gov",
    "www.{s}county.ga.gov",
    "www.{s}countyga.org",
    "www.{s}county.com",
    "www.{s}county.org",
    "www.{s}countyga.com",
    "www.co.{s}.ga.us",
    "www.{s}.ga.us",
    "{s}countyga.gov",
    "www.{s}countygeorgia.gov",
    "www.{s}county.net",
    "{s}county.net",
    "www.{s}coga.org",
    "www.{s}countyga.us",
)

# Vendor fingerprints, matched against **domains and URL paths** rather than bare
# words.
#
# Bare-substring matching does not work here and the failure is silent: "escribe"
# is contained in "describe", so every page using that ordinary English word
# reported the eSCRIBE platform. Vendors are identified by the domain they serve
# assets from, which is precise and cannot collide with prose.
PLATFORM_MARKERS = {
    r"civicplus\.com|/agendacenter": "CivicPlus",
    r"granicus\.com": "Granicus",
    r"legistar\.com": "Legistar",
    r"civicclerk\.com": "CivicClerk",
    r"boarddocs\.com": "BoardDocs",
    r"novusagenda\.com": "NovusAGENDA",
    r"iqm2\.com": "IQM2",
    r"civicweb\.net": "CivicWeb",
    r"swagit\.com": "Swagit",
    r"agendaquick\.com": "AgendaQuick",
    r"primegov\.com": "PrimeGov",
    r"escribemeetings\.com": "eSCRIBE",
    r"\bonbase\b": "OnBase",
}

# Municode hosts municipal *codes*, not agendas. It shows up on most county sites
# and would inflate the "platform identified" count while buying no minutes, so it
# is tracked separately rather than counted as an agenda platform.
NON_AGENDA_MARKERS = {r"municode\.com": "Municode (code library, not agendas)"}

_PLATFORM_RE = {re.compile(p, re.I): name for p, name in PLATFORM_MARKERS.items()}
_NON_AGENDA_RE = {re.compile(p, re.I): name for p, name in NON_AGENDA_MARKERS.items()}

# Counties whose government does not follow the usual hostname patterns, most
# because of a consolidated city-county government. Explicit, so a miss is a
# fixable table entry rather than a silent gap.
SITE_OVERRIDES = {
    "Richmond": "https://www.augustaga.gov/",       # consolidated Augusta-Richmond
    "Clarke": "https://www.accgov.com/",            # consolidated Athens-Clarke
    "Bulloch": "https://bullochcounty.net/",
    "Monroe": "https://www.monroecoga.org/",
    "Emanuel": "https://www.emanuelco-ga.gov/",
    "Cook": "https://www.cookcountyga.us/",
    "Pike": "https://www.pikecoga.com/",
    "Twiggs": "https://twiggscountyga.gov/",
    "Wilkes": "https://www.wilkescountyga.gov/",
}

# Link text that suggests the page behind it lists agendas or minutes.
_MINUTES_LINK_RE = re.compile(
    r"agenda|minutes|board of commissioners|county commission|zoning|planning commission",
    re.I,
)
_LINK_RE = re.compile(r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', re.I | re.S)
_TAG_RE = re.compile(r"<[^>]+>")


@dataclass
class CountyRecon:
    """What was found for one county."""

    county: str                       # tracker form
    site_url: str | None = None
    platforms: list[str] = field(default_factory=list)
    minutes_urls: list[str] = field(default_factory=list)
    other_vendors: list[str] = field(default_factory=list)
    note: str = ""

    @property
    def short(self) -> str:
        return self.county.replace(" County, Georgia", "")

    @property
    def status(self) -> str:
        if not self.site_url:
            return "site not found"
        if self.platforms:
            return "platform identified"
        return "site found, platform unknown"


def _get(url: str, *, timeout: int = 10) -> requests.Response | None:
    try:
        r = requests.get(
            url, headers={"User-Agent": USER_AGENT}, timeout=timeout, allow_redirects=True
        )
        return r if r.status_code < 400 else None
    except requests.RequestException:
        return None


def detect_platforms(*sources: str) -> list[str]:
    """Agenda-platform vendors visible in any of the given strings (markup, URLs)."""
    hay = " ".join(s for s in sources if s)
    return sorted({name for rx, name in _PLATFORM_RE.items() if rx.search(hay)})


def detect_non_agenda(*sources: str) -> list[str]:
    """Vendors present but irrelevant to minutes, reported so they are not mistaken
    for an agenda system."""
    hay = " ".join(s for s in sources if s)
    return sorted({name for rx, name in _NON_AGENDA_RE.items() if rx.search(hay)})


def find_site(county_short: str) -> str | None:
    """Probe candidate hostnames for a county's official website."""
    if county_short in SITE_OVERRIDES:
        response = _get(SITE_OVERRIDES[county_short], timeout=12)
        if response is not None:
            return response.url
    slug = county_short.lower().replace(" ", "")
    for pattern in HOST_PATTERNS:
        response = _get(f"https://{pattern.format(s=slug)}", timeout=8)
        # A real county homepage, not a parked domain or an error page.
        if response is not None and len(response.text) > 800:
            return response.url
    return None


def minutes_links(site_url: str, html: str, *, limit: int = 6) -> list[str]:
    """Same-site links whose text suggests agendas or minutes."""
    host = urlparse(site_url).netloc
    found: list[str] = []
    for href, label in _LINK_RE.findall(html):
        text = _TAG_RE.sub(" ", label)
        if not _MINUTES_LINK_RE.search(text):
            continue
        absolute = urljoin(site_url, href)
        parsed = urlparse(absolute)
        if parsed.scheme not in ("http", "https"):
            continue
        # Keep same-site links and known-vendor offsite links; drop the rest.
        offsite_vendor = bool(detect_platforms(absolute))
        if parsed.netloc != host and not offsite_vendor:
            continue
        if absolute not in found:
            found.append(absolute)
        if len(found) >= limit:
            break
    return found


def recon_county(county: str) -> CountyRecon:
    """Find one county's site and identify its agenda platform."""
    result = CountyRecon(county=county)
    site = find_site(result.short)
    if not site:
        result.note = "no candidate hostname answered; needs a manual lookup"
        return result
    result.site_url = site

    home = _get(site)
    home_html = home.text if home else ""
    result.platforms = detect_platforms(home_html, site)
    result.other_vendors = detect_non_agenda(home_html, site)

    # The vendor is often only visible one click in, on the agendas page.
    candidates = minutes_links(site, home_html)
    for url in candidates:
        if detect_platforms(url):
            result.minutes_urls.append(url)
            result.platforms = sorted(set(result.platforms) | set(detect_platforms(url)))
            continue
        page = _get(url)
        if page is None:
            continue
        found = detect_platforms(page.text, page.url)
        if found:
            result.minutes_urls.append(page.url)
            result.platforms = sorted(set(result.platforms) | set(found))
        if len(result.minutes_urls) >= 3:
            break

    if not result.platforms:
        result.note = "site found but no known vendor fingerprint; likely bespoke or PDF-only"
    return result


def target_counties(workbook: Path) -> list[str]:
    """Counties worth scraping: those with a facility or a recorded local action.

    A county with neither has nothing for this strand to find, so scraping it
    would cost 159ths of the effort for none of the signal.
    """
    from openpyxl import load_workbook

    ws = load_workbook(workbook, read_only=True)["Transformed"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [dict(zip(header, r)) for r in it]

    def num(row, key):
        v = row.get(key)
        return v if isinstance(v, (int, float)) else 0

    return sorted(
        r["county"]
        for r in rows
        if num(r, "dc_mapped_n")
        or num(r, "dc_permitted_n")
        or num(r, "dc_institutional_n")
        or num(r, "dc_local_action")
    )


def recon(counties: list[str], *, workers: int = 8, verbose: bool = True) -> list[CountyRecon]:
    results: list[CountyRecon] = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for result in pool.map(recon_county, counties):
            if verbose:
                print(f"  {result.short:<12} {result.status:<28} {', '.join(result.platforms)}")
            results.append(result)
    return sorted(results, key=lambda r: r.county)


def format_report(results: list[CountyRecon]) -> str:
    """Render the recon as a worklist, ordered by how much each adapter buys."""
    by_platform: dict[str, list[str]] = {}
    for r in results:
        for p in r.platforms:
            by_platform.setdefault(p, []).append(r.short)

    identified = [r for r in results if r.platforms]
    unknown = [r for r in results if r.site_url and not r.platforms]
    missing = [r for r in results if not r.site_url]

    lines = [
        "County minutes recon: where each target county publishes agendas and minutes",
        "=" * 78,
        "",
        "Target set: counties with a data center or a recorded ordinance or moratorium.",
        "Counties with neither are skipped; they have nothing for this strand to find.",
        "",
        f"{len(results)} counties targeted.",
        f"  {len(identified)} with an identified agenda platform",
        f"  {len(unknown)} site found, platform not fingerprinted",
        f"  {len(missing)} site not found, needs a manual lookup",
        "",
        "PLATFORMS, by how many target counties each adapter would cover",
        "-" * 78,
    ]
    for platform, counties in sorted(by_platform.items(), key=lambda kv: -len(kv[1])):
        lines.append(f"  {platform:<14} {len(counties):>2}  {', '.join(sorted(counties))}")

    lines += ["", "PER COUNTY", "-" * 78]
    for r in results:
        lines.append(f"{r.short}")
        lines.append(f"  site      {r.site_url or '(not found)'}")
        if r.platforms:
            lines.append(f"  platform  {', '.join(r.platforms)}")
        for u in r.minutes_urls:
            lines.append(f"  minutes   {u}")
        if r.other_vendors:
            lines.append(f"  other     {', '.join(r.other_vendors)}")
        if r.note:
            lines.append(f"  note      {r.note}")
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Find where each target county publishes commission and zoning minutes."
    )
    parser.add_argument("--workbook", type=Path, default=Path("data/processed/ga_data_centers.xlsx"))
    parser.add_argument("--out", type=Path, default=Path("data/interim/minutes-recon.txt"))
    args = parser.parse_args()

    counties = target_counties(args.workbook)
    print(f"Recon on {len(counties)} target counties...")
    results = recon(counties)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(format_report(results))
    print(f"\n-> {args.out}")


if __name__ == "__main__":
    main()
